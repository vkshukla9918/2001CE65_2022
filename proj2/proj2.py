import time
from datetime import datetime
start_time = datetime.now()
import streamlit as st
import numpy as np
from numpy import nan
import pandas as pd
from pathlib import Path
import math 
import openpyxl                          #imported library
from openpyxl import workbook
from openpyxl.styles import Border,Side
from openpyxl.styles import PatternFill


def octact_identification(mod):
	data.at[2,' '] = 'Mod ' + str(mod)   
	
	division = math.ceil(len(data)/mod)
	
	#for loop for printing the range division
	for i in range(division):

		if((i+1)*mod-1) < len(data):
			data.at[i+3,'Overall Octant Count'] = str((i)*mod)+'-'+str((i+1)*mod-1)
			for j, x in enumerate(['1','-1','2','-2','3','-3','4','-4']):
				data.at[i+3,' '*(j+2)] = data['octant'].iloc[(i)*mod :(i+1)*mod].value_counts()[int(x)]
		else:
			data.at[i+3,'Overall Octant Count'] = str((i)*mod)+'-'+str(len(data)-1)         
		#using iloc and value_counts for count of octant in the dicided range.    
			for j, x in enumerate(['1','-1','2','-2','3','-3','4','-4']):
				data.at[i+3,' '*(j+2)] = data['octant'].iloc[(i)*mod :len(data)].value_counts()[int(x)]
		


try:
	st.title('Welcome to Octant Batch Processing')
	uploaded_file = st.file_uploader("Choose a file",type=["xlsx"], accept_multiple_files=True)
	Mod = int(st.number_input("Enter the mod value:"))
except:
	print('strealit is not imported')
b=st.button('Compute')

if Mod != 0 &b:
	st.success('Done')
	
	if uploaded_file is not None:
		for f in uploaded_file:                #loop for iteration on each file after completion
			
			filename = f.name
			filename = filename.split('.xlsx')[0]

			data = pd.read_excel(f)
			# data = data.head(200)

			# data = data.head(1000)
			data.at[0,'U_avg'] =  round(data['U'].mean(),3)
			data.at[0,'V_avg'] = round(data['V'].mean(),3)
			data.at[0,'W_avg'] = round(data['W'].mean(),3)
			data['U-U_avg'] = round(data['U'] - data['U_avg'][0],3)
			data['V-V_avg'] = round(data['V'] - data['V_avg'][0],3)
			data['W-W_avg'] = round(data['W'] - data['W_avg'][0],3)
			
			octant = []
			try:
				for i in range(len(data)):
					#taking new variables for acessing each element of 'U','V' and 'W' Columns
					x = data['U-U_avg'][i]
					y = data['V-V_avg'][i]
					z = data['W-W_avg'][i]
				#octant identification and adding to octact list,octant value of each row.
					if x >= 0 and y >= 0 and z >= 0:
						octant.append(1)
					elif x >= 0 and y >= 0 and z < 0:
						octant.append(-1)
					elif x < 0 and y >= 0 and z >= 0:
						octant.append(2)
					elif x < 0 and y >= 0 and z < 0:
						octant.append(-2)
					elif x < 0 and y < 0 and z >= 0:
						octant.append(3)
					elif x < 0 and y < 0 and z < 0:
						octant.append(-3)
					elif x >= 0 and y < 0 and z >= 0:
						octant.append(4)
					elif x >= 0 and y < 0 and z < 0:
						octant.append(-4)
			except:
				print('error1')
			data['octant'] = octant.copy()   #creted new column octant and stored the value of list octant
			data[''] = ' '
			data[' '] = ' '
			data.at[1,'Overall Octant Count'] = 'Octant ID'
			data.at[2,'Overall Octant Count'] = 'Overall Count'



			for j, x in enumerate(['1','-1','2','-2','3','-3','4','-4']):
				data.at[1,' '*(j+2)] = x
				data.at[2,' '*(j+2)] = octant.count(int(x))      #counting the values of each octant and storing the value of overall count
			try:
				mod = Mod  #it decides in which range i have to break the data
				
				y = math.ceil(len(data)/mod)
				octact_identification(mod)        #function for range division and corresponding to that the value of count of an octant
			except:
				print('mod is not defined')
				


			overall_count = [octant.count(1),octant.count(-1),octant.count(2),octant.count(-2),octant.count(3),octant.count(-3),octant.count(4),octant.count(-4)]
			overall_count.sort(reverse = True)
			for j, x in enumerate(['1','-1','2','-2','3','-3','4','-4']):
				data.at[1,' '*(j+10)] = 'Rank of ' + x
			data.at[5+y,' '*17] = 'Internal outward interaction'
			data.at[6+y,' '*17] = 'External outward interaction'
			data.at[7+y,' '*17] = 'External Ejection'
			data.at[8+y,' '*17] = 'Internal Ejection'
			data.at[9+y,' '*17] = 'External inward interaction'
			data.at[10+y,' '*17] = 'Internal inward interaction'
			data.at[11+y,' '*17] = 'Internal sweep'
			data.at[12+y,' '*17] = 'External sweep'
			data.at[4+y,' '*18] = 'Count of Rank 1 Mod Values'
			data.at[1,' '*18] ='Rank1 Octant ID'
			data.at[4+y,' '*16] = 'Octant ID'
			data.at[4+y,' '*17] = 'Octant Name'
			data.at[1,' '*19] = 'Rank1 Octant Name'
			for j, x in enumerate(['1','-1','2','-2','3','-3','4','-4']):
				data.at[2,' '*(j+10)] = overall_count.index(octant.count(int(x))) + 1
				if(overall_count[0] == octant.count(int(x))):
					data.at[2,' '*18]= x
					if(data.at[2,' '*18] == x):                             #ranking of octant
						data.at[2,' '*19] = data.at[5+y+j, ' '*17]
			

			list = []
			
			for i in range(y):
				mod_count = [data['octant'].iloc[(i)*mod :(i+1)*mod].value_counts()[1],data['octant'].iloc[(i)*mod :(i+1)*mod].value_counts()[-1],data['octant'].iloc[(i)*mod :(i+1)*mod].value_counts()[2],data['octant'].iloc[(i)*mod :(i+1)*mod].value_counts()[-2],data['octant'].iloc[(i)*mod :(i+1)*mod].value_counts()[3],data['octant'].iloc[(i)*mod :(i+1)*mod].value_counts()[-3],data['octant'].iloc[(i)*mod :(i+1)*mod].value_counts()[4],data['octant'].iloc[(i)*mod :(i+1)*mod].value_counts()[-4]]
				mod_count.sort(reverse = True)
				for j, x in enumerate(['1','-1','2','-2','3','-3','4','-4']):
					data.at[3+i,' '*(j+10)] = mod_count.index(data['octant'].iloc[(i)*mod :(i+1)*mod].value_counts()[int(x)]) + 1
					if(mod_count[0] == data['octant'].iloc[(i)*mod :(i+1)*mod].value_counts()[int(x)]):
						data.at[3+i,' '*18]= x
						list.append(int(x))
						if(data.at[3+i,' '*18] == x):
							data.at[3+i,' '*19] = data.at[5+y+j, ' '*17]
			

			for i,x in enumerate(['1','-1','2','-2','3','-3','4','-4']):
				data.at[5+y+i,' '*18] = list.count(int(x))
				data.at[5+i+y,' '*16] = x
			data[' '*20] = ' '
			data.at[2,' '*21] = 'From'
			data.at[1,'Overall transition Count'] = 'Octant#'
			
			try:
				for i, x in enumerate(['1','-1','2','-2','3','-3','4','-4']):
					data.at[2+i,'Overall transition Count'] = x
					data.at[1,' '*(22+i)] = x
				dict = {1:{1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0},
				-1:{1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0},
				2:{1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0},
				-2:{1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0},
				3:{1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0},   #created nested dictionaryfor the value of transition count initial at 0
				-3:{1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0},
				4:{1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0},
				-4:{1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}}
					
				for i in range(len(data)-1):
					dict[octant[i]][octant[i+1]] += 1
				for j, x in enumerate(['1','-1','2','-2','3','-3','4','-4']):
					data.at[2,' '*(22+j)] = dict[1][int(x)]
					data.at[3,' '*(22+j)] = dict[-1][int(x)]
					data.at[4,' '*(22+j)] =dict[2][int(x)]
					data.at[5,' '*(22+j)] =dict[-2][int(x)]        #filling the trasition count in table using loop
					data.at[6,' '*(22+j)] =dict[3][int(x)]
					data.at[7,' '*(22+j)] =dict[-3][int(x)]
					data.at[8,' '*(22+j)] =dict[4][int(x)]
					data.at[9,' '*(22+j)] =dict[-4][int(x)]
					data.style.highlight_max(color = 'Yellow', axis = 0)
			except:
				print('error3')
			
			try:
				def octant_transition_count(mod):          #creating new function for mod transition count
					dict_mod = {1:{1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0},
					-1:{1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0},
					2:{1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0},
					-2:{1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0},
					3:{1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0},
					-3:{1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0},
					4:{1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0},
					-4:{1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}}
					for i in range(y):
						data.at[13+i*13,'Overall transition Count']='Mod Transition Count'
						data.at[14+i*13,' '*22] = 'To'
						data.at[15+i*13,'Overall transition Count'] = 'octant#'    #columns for mod transition count
						data.at[16+i*13,' '*21]= 'From'
						if((i+1)*mod-1) < len(data)-1:
							data.at[14+i*13,'Overall transition Count'] = str((i)*mod)+'-'+str((i+1)*mod-1)
							for j in range(i*mod,(i+1)*mod):
								dict_mod[octant[j]][octant[j+1]] += 1
							for j, x in enumerate(['1','-1','2','-2','3','-3','4','-4']):
								data.at[16+i*13,' '*(22+j)] = dict_mod[1][int(x)]
								data.at[17+i*13,' '*(22+j)] = dict_mod[-1][int(x)]
								data.at[18+i*13,' '*(22+j)] =dict_mod[2][int(x)]
								data.at[19+i*13,' '*(22+j)] =dict_mod[-2][int(x)]
								data.at[20+i*13,' '*(22+j)] =dict_mod[3][int(x)]
								data.at[21+i*13,' '*(22+j)] =dict_mod[-3][int(x)]      #fillimg the table upto second last range for transition count
								data.at[22+i*13,' '*(22+j)] =dict_mod[4][int(x)]
								data.at[23+i*13,' '*(22+j)] =dict_mod[-4][int(x)]
								
								dict_mod[1][int(x)] = 0
								dict_mod[-1][int(x)] = 0
								dict_mod[2][int(x)] = 0
								dict_mod[-2][int(x)] = 0  #making here again all  0
								dict_mod[3][int(x)] = 0
								dict_mod[-3][int(x)] = 0
								dict_mod[4][int(x)] = 0
								dict_mod[-4][int(x)] = 0

						else:
							data.at[14+i*13,'Overall transition Count'] = str((i)*mod)+'-'+str(len(data)-1)
							for j in range(i*mod,len(data)-1):
								dict_mod[octant[j]][octant[j+1]] += 1
							for j, x in enumerate(['1','-1','2','-2','3','-3','4','-4']):
								data.at[16+i*13,' '*(22+j)] = dict_mod[1][int(x)]
								data.at[17+i*13,' '*(22+j)] = dict_mod[-1][int(x)]
								data.at[18+i*13,' '*(22+j)] =dict_mod[2][int(x)]
								data.at[19+i*13,' '*(22+j)] =dict_mod[-2][int(x)]
								data.at[20+i*13,' '*(22+j)] =dict_mod[3][int(x)]     #for lat range which would not be of 5000 length
								data.at[21+i*13,' '*(22+j)] =dict_mod[-3][int(x)]
								data.at[22+i*13,' '*(22+j)] =dict_mod[4][int(x)]
								data.at[23+i*13,' '*(22+j)] =dict_mod[-4][int(x)]
						for j,x in enumerate(['1','-1','2','-2','3','-3','4','-4']):
							data.at[16+j+i*13,'Overall transition Count'] = x
							data.at[15+i*13,' '*(22+j)] = x
			except:
				print('error found in function octant_transition_count')
				
				
				
			octant_transition_count(mod)

			total_count = 0
			data[' '*30] = ' '
			data.at[1,'Longest Subsquence Length'] = 'octant##'
			data.at[1,' '*31] = 'Longest Subsquence Length'
			data.at[1,' '*32] = 'Count'
			longest_count = {1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
			try:
				for j, x in enumerate(['1','-1','2','-2','3','-3','4','-4']):
					temp = 0
					for i in range(len(data)-1):
						if(data.at[i,'octant'] == data.at[i+1,'octant'] and data.at[i,'octant'] == int(x)):
							temp += 1
							if(i == len(data)-2):
								longest_count[int(x)] = max(longest_count[int(x)],temp+1)

						
						else:
							longest_count[int(x)] = max(longest_count[int(x)],temp+1)
							temp = 0
			except:
				print('error4')
			

			longestoccur_count = {1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
			try:
				for j, x in enumerate(['1','-1','2','-2','3','-3','4','-4']):
					temp = 0
					for i in range(len(data)-1):
						if(data.at[i,'octant'] == data.at[i+1,'octant'] and data.at[i,'octant'] == int(x)):
							temp += 1
							if(i == len(data)-2):
								if(temp+1 == longest_count[int(x)]):
									longestoccur_count[int(x)] += 1

						else:
							if(temp+1 == longest_count[int(x)]):
								longestoccur_count[int(x)] += 1
							temp = 0
			except:
				print('octant is not defined')

			for j,x in enumerate(['1','-1','2','-2','3','-3','4','-4']):
				data.at[2+j,'Longest Subsquence Length'] = x
				data.at[2+j,' '*31]= longest_count[int(x)]
				data.at[2+j,' '*32] = longestoccur_count[int(x)]
				total_count += longestoccur_count[int(x)]
			

			data[' '*33] = ' '
			
			
			data.at[1,'longest Subsquence Length with range'] = 'octant##'
			data.at[1,' '*34] = 'Longest Subsquence Length'
			data.at[1,' '*35] = 'Count'
			row  = 2
			for j,x in enumerate(['1','-1','2','-2','3','-3','4','-4']):
				data.at[row,'longest Subsquence Length with range' ] = x
				data.at[row, ' '*34] = longest_count[int(x)]
				data.at[row,' '*35] = longestoccur_count[int(x)]

				data.at[row+1, 'longest Subsquence Length with range'] = 'Time'
				data.at[row+1, ' '*34] = 'From'
				data.at[row+1, ' '*35] = 'To'
				row += 2
				for i in range(len(data)):
					if data.at[i, 'octant'] == int(x):
						initial = i #storing the row of first time range
						occurance = 0
						while(data.at[i, 'octant'] == int(x)):
							i += 1
							occurance += 1
							if(i == len(data)):
								i -= 1     #if i becomes equal to length of data then reducing it by 1 as row will start from 0.
								break
						if occurance == longest_count[int(x)]: # if longestsubsequence length is matched then print that range of time 
							data.at[row, ' '*34] = data.at[initial, 'T']
							data.at[row, ' '*35] = data.at[i, 'T']
							row += 1 #shifting to next row
						occurance = 0 #making occurance again 0 to search next longest subsequence length and their range
			data.replace(nan,'',inplace=True) #removing nan
			timestr = time.strftime("%Y-%m-%d-%H-%M-%S")

			data.to_excel(filename + '_mod'+str(mod) +'_' + timestr +'.xlsx',index = False)   #printing to file without border and colour
			try:
				wb = openpyxl.load_workbook(filename + '_mod'+str(mod) +'_' + timestr +'.xlsx')  #opened file using openpyxl
				ws = wb['Sheet1'] #active sheet
			except:
				print('file is not found')
			try:
				top = Side(border_style = 'thin', color = '000000')
				bottom = Side(border_style = 'thin', color = '000000')  #border specification
				left = Side(border_style = 'thin', color = '000000')
				right = Side(border_style = 'thin', color = '000000')
				border = Border(top = top, bottom = bottom, left = left, right = right)
			except:
				print('Border or side is not imported')

			fill_pattern = PatternFill( start_color ='FFFF00',end_color ='FFFF00', fill_type='solid' )  #colour fill specification
			try:
				for i in range(3, y+5):
					for j in range(14,33):
						ws.cell(row = i ,column = j).border = border    #border for table as per sample output file

				for i in range(4,y+5):
					for j in range(23,31):   
						if ws.cell(row =i,column=j).value == 1:              #colour which cell value is 1
							ws.cell(row = i ,column = j).fill = fill_pattern

				for i in range(y+6, y+15):
					for j in range(29,32):
						ws.cell(row = i ,column = j).border = border
				for i in range(3, y+8):
					for j in range(35,44):
						ws.cell(row = i ,column = j).border = border
				a =[] #creted list to store max value in each row
				for i in range(4, 12):
					for j in range(36,44):
						a.append(ws.cell(row = i ,column = j).value)
					z = max(a)                                              #colour to max value cell
					for j in range(36,44):
						if ws.cell(row = i ,column = j).value == z:
							ws.cell(row = i ,column = j).fill = fill_pattern
					a = []  #making it again 0
			
					
					
				for i in range(y):
					for j in range(17+i*13, 26+i*13):
						for k in range(35,44):
							ws.cell(row = j ,column = k).border = border
				b = []#creted list to store max value in each row
				for i in range(y):
					for j in range(18+i*13, 26+i*13):
						for k in range(36,44):
							b.append(ws.cell(row = j ,column = k).value)     #colour to max value cell
						z = max(b)
						for k in range(36,44):
							if ws.cell(row = j ,column = k).value == z:
								ws.cell(row = j ,column = k).fill = fill_pattern
						b = []   #making it again 0
			


				for i in range(3, 12):
					for j in range(45,48):
						ws.cell(row = i ,column = j).border = border
				for i in range(3, 20+total_count):
					for j in range(49,52):
						ws.cell(row = i ,column = j).border = border
			except:
				print('error5')

			
			wb.save(filename + '_mod'+str(mod) +'_' + timestr +'.xlsx')
			
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
